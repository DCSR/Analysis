
"""
Steven:
$pip install openpyxl

See: https://www.youtube.com/watch?v=xHoXFWgmB84

"""
from openpyxl import Workbook
import ListLib


def pushInjTimesToExcel(aRecordList):

    startRow = 2
    startColumn = 2

    myWorkBook = Workbook()
    sheet = myWorkBook.active
    sheet.title = 'SEAWAKE Data'



    for aRecord in aRecordList:
        if aRecord.fileName != "empty": 
            injections = ['Inj']
            durations = ['Duration']
            secTimes = ['Time(sec)']
            minTimes = ['Time(min)']
            intervals = ['Interval(sec)']

            my2DList = [injections,durations,secTimes,minTimes,intervals]
            injection = 0
            previousInjTime = 0
            print("Inj Duration   Time (sec)   Time (min) Interval (sec)\n")
            pumpOn = False
            for pairs in aRecord.datalist:
                if pairs[1] == 'P':
                    pumpStartTime = pairs[0]
                    injection = injection + 1
                    injections.append(injection)
                    secTime = pairs[0]/1000
                    secTimes.append(secTime)
                    minTime = secTime/60
                    minTimes.append(minTime)
                    interval = secTime - previousInjTime
                    intervals.append(interval)
                    previousInjTime = secTime
                    pumpOn = True
                if pairs[1] == 'p':
                    if pumpOn:
                        pumpOn = False
                        duration = pairs[0]-pumpStartTime
                        durations.append(duration)
                        if injection == 1:
                            tempString = "{0} {1:10.2f} {2:10.2f} {3:10.2f}".format(injection,duration,secTime,minTime,interval)
                        else:
                            tempString = "{0} {1:10.2f} {2:10.2f} {3:10.2f} {4:10.2f}".format(injection,duration,secTime,minTime,interval)
                        #print(tempString)
                    
            print("Number of injections: "+str(injection)+"\n")
            print(my2DList)

            # ************ openpyxl stuff **************

            categories = len(my2DList)
            dataLength = len(my2DList[0])

            cellID = sheet.cell(row=1,column=startColumn+2)
            cellID.value = aRecord.fileName

            for d in range(0,dataLength):
                for c in range(0,categories):    
                    cellID = sheet.cell(row=startRow+d,column=startColumn+c) # The list is zero based, sheet is not
                    cellID.value = my2DList[c][d]

            startColumn = sheet.max_column + 2

    print('WorkSheet size =', sheet.max_row,'x',sheet.max_column)

    myWorkBook.save(filename="ExcelTest.xlsx")


            #**** work on push TH to excel here ****

def pushTHToExcel(aRecordList):

    startRow = 2
    startColumn = 2

    myWorkBook = Workbook()
    sheet = myWorkBook.active
    sheet.title = 'TH Data'


    for aRecord in aRecordList:
        if aRecord.fileName != "empty":
            injPerBlock = ['Inj per block']
##            injections = ['Injections']
##            blockCount = ['Blocks']
##            responses = ['Responses']

            
            my2DList = [injPerBlock]

            aList = aRecord.datalist

            pump_count_list = ListLib.get_pump_count_per_block(aList)
            for item in pump_count_list:
                injPerBlock.append(item)
                print(str(item))


            
                #*** this code works so long as injPerBlock is not used
                #*** use for another TH button or find a way to integrate?

##            aList = aRecord.datalist
##            
##            count = ListLib.count_char('L',aList)
##            responses.append(count)
##            print(responses)
##            
##            count = ListLib.count_char('P',aList)
##            injections.append(count)

##    
##            count = ListLib.count_char('B',aList)
##            blockCount.append(count)
            
##             ****working inj per block****
##            pump_count_list = ListLib.get_pump_count_per_block(aList)
##            for item in pump_count_list:
##                injPerBlock.append(item)
##                print(str(item))



            print(my2DList)

                # ************ openpyxl stuff **************

            categories = len(my2DList)
            dataLength = len(my2DList[0])

            cellID = sheet.cell(row=1,column=startColumn)
            cellID.value = aRecord.fileName



            for d in range(0,dataLength):
                for c in range(0,categories):    
                    cellID = sheet.cell(row=startRow+d,column=startColumn+c) # The list is zero based, sheet is not
                    cellID.value = my2DList[c][d] #this only seems to work when a lists are the same length, figure out
                                                    #a way to do this with lists of different lengths

            startColumn = sheet.max_column + 1

        print('WorkSheet size =', sheet.max_row,'x',sheet.max_column)

        myWorkBook.save(filename="ExcelTest.xlsx")


   
    

          
